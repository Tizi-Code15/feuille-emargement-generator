from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
from datetime import datetime, timedelta

import locale

def generate_excel(promotion_name, start_date, stagiaires):
    # Configurer la locale en français
    locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')

    # Charger le fichier Excel existant depuis le dossier 'feuille'
    file_path = os.path.join("feuille", "classeur.xlsx")  # Charger le fichier "classeur.xlsx" depuis le dossier "feuille"
    wb = load_workbook(file_path)

    # Convertir start_date en format datetime
    start_date_obj = datetime.strptime(start_date, "%Y-%m-%d")

    # Ajouter des images depuis le dossier 'logo'
    image1_path = os.path.join("logo", "M.jpg")  # Chemin vers l'image 1
    image2_path = os.path.join("logo", "G.jpg")  # Chemin vers l'image 2

    # Vérifier si les images existent
    if os.path.exists(image1_path) and os.path.exists(image2_path):
        # Charger les images
        img1 = Image(image1_path)
        img2 = Image(image2_path)
    else:
        print("Une ou plusieurs images sont manquantes!")

    # Diviser les stagiaires en groupes de 10
    stagiaires_per_page = 10
    num_pages = (len(stagiaires) + stagiaires_per_page - 1) // stagiaires_per_page  # Calcul du nombre de pages nécessaires

    # Debug: Affichage du nombre de stagiaires et de pages
    print(f"Nombre de stagiaires : {len(stagiaires)}")
    print(f"Nombre de pages nécessaires : {num_pages}")

    # Pour chaque page, ajouter une nouvelle feuille et insérer les stagiaires
    for page_num in range(num_pages):
        # Créer une nouvelle feuille pour chaque page sauf la première
        if page_num == 0:
            ws = wb.active  # La première page utilise la feuille active
            ws.title = f"Page 1"  # Donner un titre à la première feuille
        else:
            ws = wb.create_sheet(title=f"Page {page_num + 1}")  # Créer une nouvelle feuille pour les pages suivantes

        # Insérer le nom de la promotion et la date dans la feuille
        ws['E2'] = f"Promotion: {promotion_name}"  # Insertion du nom de la promotion
        ws['E3'] = f"Date: {start_date}"  # Insertion de la date

        # Afficher les dates de la semaine dans les cellules spécifiées
        date_cells = ['C5', 'E5', 'G5', 'I5', 'K5']  # Modification des cellules où afficher les dates

        for i, cell in enumerate(date_cells):
            # Calculer la date de chaque jour de la semaine
            day_date = start_date_obj + timedelta(days=i)
            # Insérer la date du jour au format "Jour 10/03/2025"
            ws[cell] = day_date.strftime("%A %d/%m/%Y")  # Format de la date en français

        # Insérer les stagiaires à partir de la cellule spécifiée
        start_row = 8  # La première ligne où insérer les stagiaires
        start_col = 2  # Colonne B (2) pour les noms et prénoms

        # Debug: Affichage des stagiaires insérés sur chaque page
        page_stagiaires = stagiaires[page_num * stagiaires_per_page: (page_num + 1) * stagiaires_per_page]
        print(f"Stagiaires pour la page {page_num + 1}: {page_stagiaires}")
        
        # Insérer les stagiaires dans la feuille
        row_num = start_row
        for stagiaire in page_stagiaires:
            ws.cell(row=row_num, column=start_col, value=f"{stagiaire[0]} {stagiaire[1]}")  # Nom et Prénom
            row_num += 1

        # Ajouter les images dans la feuille
        if page_num == 0:  # Les images ne sont ajoutées qu'à la première page
            img1.anchor = 'B20'  # Ancrage de l'image dans la cellule B20
            img2.anchor = 'H20'  # Ancrage de l'image dans la cellule H20
            ws.add_image(img1)
            ws.add_image(img2)

    # Chemin pour le fichier généré dans le dossier parent
    generated_files_dir = os.path.join(os.getcwd(), "generated_files")

    # Vérifier si le dossier 'generated_files' existe, sinon le créer
    if not os.path.exists(generated_files_dir):
        os.makedirs(generated_files_dir)

    # Chemin du fichier Excel modifié
    modified_file_path = os.path.join(generated_files_dir, f"feuille_modifiee_{promotion_name}_{start_date}.xlsx")

    # Sauvegarder les modifications dans le fichier Excel
    wb.save(modified_file_path)

    return modified_file_path
